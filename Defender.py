import datetime
import tempfile
import urllib3
import calendar
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

strings = list()
result_string = list()
max_string_size = 0
max_string_value = ""

class User:
    detection_string = ""
    ip = ""

class Detections:
    ips = ""
    Users = None
    Wacatac = 0
    Bearfoos = 0
    total = 0


def as_text(value):
    if value is None:
        return ""
    return str(value)


def remove_trash_in_string(source):
    result = ""
    all = source.split("[]")
    if len(all) >= 1:
        result = all[len(all) - 1]
    else:
        result = source
    return result


def save_report_to_exel(reportdate, dest_filename, detections_by_hours):
    # create workbook
    wb = Workbook()
    # get active worksheet
    ws1 = wb.active
    ws1.title = "report"
    # create the header or the table
    ws1.merge_cells('B2:E2')
    ws1['A2'] = "   "
    ws1['B2'] = reportdate
    ws1['C3'] = "All detections"
    ws1['D3'] = "Trojan:Win32/Wacatac"
    ws1['E3'] = "Trojan:Win32/Bearfoos"
    ws1['F3'] = "IP"
    ws1['G3'] = "the detections strings"

    al = Alignment(horizontal="center", vertical="center")
    fill = PatternFill("solid", fgColor="FF8080")

    total = 0
    Wacatac = 0
    Bearfoos = 0

    current_row = 4
    for hour in range(0, 24):
        ws1["B" + str(current_row)] = str(hour) + " hour"
        if hour in detections_by_hours:
            detection = detections_by_hours[hour]
        else:
            detection = Detections()
            detection.Users = dict()

        ws1["C" + str(current_row)] = detection.total
        ws1["D" + str(current_row)] = detection.Wacatac
        ws1["E" + str(current_row)] = detection.Bearfoos

        total = total + detection.total
        Wacatac = Wacatac + detection.Wacatac
        Bearfoos = Bearfoos + detection.Bearfoos

        # set color for warning hours
        if detection.total >= 10:
            ws1["B" + str(current_row)].fill = fill
            ws1["C" + str(current_row)].fill = fill
            ws1["D" + str(current_row)].fill = fill
            ws1["E" + str(current_row)].fill = fill

        if len(detection.Users) > 0:
            for ip in detection.Users:
                ws1["G" + str(current_row)] = detection.Users[ip].detection_string
                print(hour,"-",ip)
                ws1["f" + str(current_row)] = detection.Users[ip].ip
                current_row = current_row + 1

        if len(detection.Users) == 0:
            current_row = current_row + 1
    ws1["B" + str(current_row)] = "all day"
    ws1["C" + str(current_row)] = total
    ws1["D" + str(current_row)] = Wacatac
    ws1["E" + str(current_row)] = Bearfoos

    ws1.column_dimensions["A"].width = 5
    ws1.column_dimensions["B"].width = 10
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 25
    ws1.column_dimensions["E"].width = 25
    ws1.column_dimensions["F"].width = 20
    ws1.column_dimensions["G"].width = 200



    rows = ws1['B2:F' + str(current_row)]
    for row in rows:
        for cell in row:
            cell.alignment = al

    wb.save(dest_filename)


def decodeString(source):
    result = ""
    for index in range(0, len(source)):
        ch = source[index]
        if ord(ch) > 128:
            ch = '*'
        if ord(ch) < 32:
            ch = ' '
        result = result + ch
    return result


def prepareFile(path):
    fh = open(path, 'rb')
    out = bytearray()
    ba = bytearray(fh.read())
    fh.close()
    for byte in ba:
        if ((byte != 10) & (byte != 13)):
            if (byte < 0x20) or (byte > 0x7F):
                byte = 0x20
        out.append(byte)
    os.remove(path)
    fo = open(path, 'wb')
    fo.write(out)


def checkDefender(year, month, day, output_folder, download):
    global max_string_size
    global max_string_value
    result_string.clear()
    strings.clear()
    url = "http://172.245.127.190/write/PDATAPOST4{:04d}{:d}{:d}.txt".format(year, month, day)

    if (download == False):
        # read file
        path = "D:\\" + url.split('/')[-1]
    else:
        # download file
        path = tempfile.gettempdir() + "\\" + url.split('/')[-1]

        http = urllib3.PoolManager()
        response = http.request('GET', url)

        meta = response.info()
        file_size = int(meta.getheaders("Content-Length")[0])
        print("Downloading: %s Bytes: %s" % (path, file_size))
        f = open(path, 'wb')
        f.write(response.data)
        f.close()

    string = ""
    prepareFile(path)
    # read file
    with open(path, "r") as input:
        lines = input.readlines()
        # merge to records
        for line in lines:
            line = line.rstrip()
            if line.find("v") == 0:
                if len(string) == 0:
                    string = line
                else:
                    strings.append(string)
                    string = line
            else:
                string += line
    # remove string with errors only
    for string in strings:
        # work with regular and uac users
        # uac
        #if (string.find("v67 ") >= 0) or (string.find("v69 ") >= 0):
        if (string.find("v58 ") >= 0) or (string.find("v59 ") >= 0)  or (string.find("v69 ") >= 0) or (string.find("v62 ") >= 0) or (string.find("v67 ") >= 0) or (string.find("v65 ") >= 0):
            if string.find("[clean]") == -1:
                if string.find("Win32") > 0:
                    # if string.find("PUA") > 0:
                    string = decodeString(string)
                    if max_string_size < len(string):
                        max_string_value = string
                    max_string_size = max(max_string_size, len(string))

                    result_string.append(string)
                    # print string
    detections_by_hours = dict()
    #
    for string in result_string:
        all_srings = string.split("-NNCDIV-")
        rawstring = all_srings[0]
        wacatac = 0
        bearfoos = 0
        detection_string = ""
        if rawstring.find("Trojan:Win32/Wacatac") > 0:
            wacatac = 1
        if rawstring.find("Trojan:Win32/Bearfoos") > 0:
            bearfoos = 1
        # else:
        detection_string = remove_trash_in_string(rawstring)

        # hash = all[0].split("=")
        time = all_srings[2]
        # for winter
        # dt = datetime.datetime.utcfromtimestamp(int(time) - 18000)
        # for summer
        dt = datetime.datetime.utcfromtimestamp(int(time) - 14400)
        hour = dt.hour
        #
        ip = all_srings[1]
        # h = hash[1]

        if hour in detections_by_hours:
            detections_by_hours[hour].ips = detections_by_hours[hour].ips + "," + "'" + ip + "'"
            detections_by_hours[hour].Wacatac = detections_by_hours[hour].Wacatac + wacatac
            detections_by_hours[hour].Bearfoos = detections_by_hours[hour].Bearfoos + bearfoos
            if detection_string:
                if ip in detections_by_hours[hour].Users:
                    detections_by_hours[hour].Users[ip].detection_string = detections_by_hours[hour].Users[ip].detection_string + detection_string
                else:
                    user = User()
                    user.detection_string = detection_string
                    user.ip = ip
                    detections_by_hours[hour].Users[ip] = user
                    detections_by_hours[hour].total = detections_by_hours[hour].total + 1

        else:
            detect = Detections()
            detect.Users = dict()
            if detection_string:
                user = User()
                user.detection_string = detection_string
                user.ip = ip
                detect.Users[ip] = user
            detect.Wacatac = wacatac
            detect.Bearfoos = bearfoos
            detect.ips = "'" + ip + "'"
            detect.total = 1
            detections_by_hours[hour] = detect
    # prepare sql request:
    sql = "select stime, ip, group_concat(av), osver, windows, h from igorpixelavwindows where stime between @DateTime1 and  @DateTime2 and rs='i' and ip in ("
    for hour in detections_by_hours:
        s = detections_by_hours[hour].ips
        sql = sql + s
        sql = sql + ","

    sql = sql[:-1]
    sql += ") group by stime, ip, osver, windows, h order by stime, ip;"

    # save to xml
    date1 = calendar.month_abbr[month]
    date1 = date1 + ",{:02d}".format(day)

    date2 = calendar.month_abbr[month]
    date2 = date2 + ".{:02d}".format(day)

    save_report_to_exel(date1, output_folder + "\\Fusion.detection.report.for." + date2 + ".xlsx", detections_by_hours)

    # print by hour:
    total = 0


#  for hour in detections_by_hours:
# print "******** ", hour, " *****"
# print "total detections:", detections_by_hours[hour].total
# total = total + detections_by_hours[hour].total
# print "Wacatac detections:", detections_by_hours[hour].Wacatac
# print "Bearfoos detections:", detections_by_hours[hour].Bearfoos
# print "the detections string"
# for str in detections_by_hours[hour].unknow:
#     print "   ", str
# print  detections_by_hours[hour].ips


# print "*****************"
# print "total detections in the day:", total


# for x in range(1, 11):
# for i in range(29, 28, -1):
checkDefender(year=2020, month=9, day=30, output_folder="E:\\Nextcloud\\Documents\\FusionDefender", download=False)
print("max string size:", max_string_size, "\r\n", max_string_value)